import openpyxl


# creates excel worksheet
wb = openpyxl.Workbook()
ws = wb.active
# these are the needed columns for the truck load invoice
ws['A1'] = 'Dates'
ws['B1'] = 'Ticket Numbers'
ws['C1'] = 'Truck Number'
ws['D1'] = 'Customer/Job'
ws['E1'] = 'Loads/Hours'
ws['F1'] = 'Rate per Load'
ws['G1'] = 'Total Amount Owed'


# initialize total amount
total_amount = 0


# ask user to enter invoice data
while True:
   # get inputs from the user
   dates = input("Enter date (e.g. 01/01/2001): ")
   ticket_numbers = input("Enter ticket numbers (separated by commas): ")
   truck_number = input("Enter truck number: ")
   customer_job = input("Enter customer/job name: ")


   # validators for 
   while True:
       try:
           loads = int(input("Enter loads/hours: "))
           break
       except ValueError:
           print("Please enter a valid number for loads/hours.")


   while True:
       try:
           rate_per_load = float(input("Enter rate per load/hour: "))
           break
       except ValueError:
           print("Please enter a valid number for rate per load/hour.")


   # calculate the total amount owed for this row
   row_total = loads * rate_per_load


   # add invoice data
   try:
       ws.append([dates, ticket_numbers, truck_number, customer_job, loads, rate_per_load, row_total])
   except:
       print("Invalid data entered. Please try again.")


   # add this row's total to the grand total
   total_amount += row_total


   # ask user if they want to add another row
   choice = input("Do you want to add another row? (ENTER y for yes, n for no) ").lower()


   # validates entry
   while choice != 'y' and choice != 'n':
       choice= input("Invalid output. Do you want to add another row? (PLEASE ENTER y for yes, n for no)")


   if choice != 'y':
       break


# grand total row gets added after all rows added
ws.append(['Grand Total: ', '', '', '', '', '', total_amount])


# data formatting, font, width
for row in ws.iter_rows(min_row=1, max_row=1):
   for cell in row:
       cell.font = openpyxl.styles.Font(bold=True)
for col in ['A', 'B', 'C', 'D']:
   ws.column_dimensions[col].width = 20
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 20


# ask for file name then saves
invoice_name= input('What would you like to name the file?: ')
# saves, gives error message if some error happens
try:
   wb.save(f'{invoice_name}.xlsx')
   print("File saved successfully!")
except Exception as e:
   print(f'Error occurred: {e}')